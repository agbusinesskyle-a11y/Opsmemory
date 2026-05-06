"""XLSX -> CSV conversion for file_drop ingest (Chunk 9 step 3).

Per Codex chunk-9-step2 STEP 3 PLAN:
  - openpyxl(read_only=True, data_only=True, keep_links=False).
  - Multi-sheet: 'Tasks' if present, else first visible non-empty
    sheet with a recognizable task header.
  - Convert selected sheet to CSV via csv.writer.
  - Reuse the existing CSV parser path unchanged.
  - Store selected_sheet + ignored_sheets in source_metadata.

defusedxml.cElementTree is monkey-patched into openpyxl's XML parser
on import so XLSX parsing is hardened against billion-laughs / XXE
(per openpyxl's PyPI security note).
"""

from __future__ import annotations

import base64
import csv
import io
import logging
from typing import Any

# Codex chunk-9-step3 fix: do NOT call defusedxml.defuse_stdlib().
# That's a process-global monkey patch with surprising side effects on
# any other XML consumer. openpyxl already auto-detects defusedxml at
# import time and routes its own XML parsing through it
# (openpyxl.DEFUSEDXML == True when defusedxml is installed). Just
# require defusedxml is importable + assert openpyxl picked it up.
import defusedxml  # noqa: F401  (presence-required side effect)
import openpyxl  # noqa: E402

if not getattr(openpyxl, "DEFUSEDXML", False):
    # openpyxl ships with DEFUSEDXML auto-set when defusedxml is
    # importable. If False here, openpyxl is parsing XLSX XML through
    # the unhardened stdlib parser (billion-laughs / XXE exposure on
    # operator-supplied workbooks). Refuse to load this module.
    raise ImportError(
        "openpyxl.DEFUSEDXML is False; XLSX parsing would use the "
        "unhardened stdlib XML parser. Install defusedxml or upgrade "
        "openpyxl."
    )

log = logging.getLogger("opsmemory.xlsx_decode")


# Headers we recognize as "this sheet contains tasks." Case-insensitive
# substring match — same alias families as file_drop_parser to keep the
# detection consistent.
_SUMMARY_HINTS = ("summary", "task", "title", "action", "todo", "to-do")


class XlsxDecodeError(Exception):
    """Raised on malformed XLSX, bad base64, oversized bytes, or
    no readable sheet. Caller maps to HTTP 4xx with a code field.
    """

    def __init__(self, code: str, detail: str):
        super().__init__(detail)
        self.code = code
        self.detail = detail


def decode_xlsx_to_csv(
    xlsx_base64: str,
    *,
    max_decoded_bytes: int = 5 * 1024 * 1024,   # 5 MiB cap
    max_csv_chars: int = 200_000,               # match FileDropIngest.file_content
) -> tuple[str, dict]:
    """Decode a base64-encoded XLSX, pick a sheet, return CSV text +
    metadata about what was chosen.

    Returns (csv_text, metadata) where metadata is:
      {
        "selected_sheet": "Tasks",
        "ignored_sheets": ["Reference", "Calc"],
        "row_count": 42,
        "col_count": 6,
        "decoded_bytes": 18432,
      }

    Raises XlsxDecodeError on:
      - bad base64
      - exceeds size cap
      - bytes don't start with ZIP magic
      - openpyxl can't open / no sheets
      - no sheet has a recognizable task summary header
      - converted CSV exceeds max_csv_chars
    """
    # ---- base64 decode ----
    try:
        raw = base64.b64decode(xlsx_base64, validate=True)
    except (ValueError, base64.binascii.Error) as exc:
        raise XlsxDecodeError("xlsx_base64_invalid",
                              f"base64 decode failed: {exc!r}")

    if len(raw) > max_decoded_bytes:
        raise XlsxDecodeError(
            "xlsx_too_large",
            f"decoded bytes {len(raw)} exceeds cap {max_decoded_bytes}",
        )
    if len(raw) < 4 or not raw.startswith(b"PK\x03\x04"):
        raise XlsxDecodeError(
            "xlsx_bad_magic",
            "decoded bytes don't start with ZIP container magic (PK\\x03\\x04)",
        )

    # ---- openpyxl open ----
    try:
        wb = openpyxl.load_workbook(
            filename=io.BytesIO(raw),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise XlsxDecodeError("xlsx_open_failed",
                              f"openpyxl couldn't open the workbook: {exc!r}")

    if not wb.sheetnames:
        raise XlsxDecodeError("xlsx_no_sheets", "workbook has no sheets")

    # ---- Pick a sheet ----
    selected_name, ignored = _select_sheet(wb)
    if not selected_name:
        raise XlsxDecodeError(
            "xlsx_no_recognized_sheet",
            ("no sheet has a recognizable task header; "
             "looked for one of: " + ", ".join(_SUMMARY_HINTS)),
        )

    sheet = wb[selected_name]

    # ---- Convert to CSV ----
    # Codex chunk-9-step3 fix: stream the cap check after each row
    # so a pathological 100k-row sheet doesn't materialize 50MB in
    # memory before being rejected. read_only=True iter_rows is
    # streaming; this completes the streaming story.
    buf = io.StringIO()
    writer = csv.writer(buf)
    row_count = 0
    col_count = 0
    for row in sheet.iter_rows(values_only=True):
        # Trim trailing all-empty cells so we don't write a wide CSV
        # with mostly empty columns.
        last = len(row) - 1
        while last >= 0 and (row[last] is None or row[last] == ""):
            last -= 1
        cells = [_cell_to_str(row[i]) for i in range(last + 1)]
        writer.writerow(cells)
        row_count += 1
        col_count = max(col_count, len(cells))
        if buf.tell() > max_csv_chars:
            raise XlsxDecodeError(
                "xlsx_csv_too_large",
                f"converted CSV exceeded cap {max_csv_chars} chars at "
                f"row {row_count}; reduce sheet rows or split the file",
            )

    csv_text = buf.getvalue()

    metadata = {
        "selected_sheet": selected_name,
        "ignored_sheets": ignored,
        "row_count": row_count,
        "col_count": col_count,
        "decoded_bytes": len(raw),
    }
    log.info("xlsx_decoded", extra=metadata)
    return csv_text, metadata


def _select_sheet(wb) -> tuple[str | None, list[str]]:
    """Pick the sheet to convert. 'Tasks' wins if present and visible.
    Otherwise the first visible non-empty sheet whose first row
    contains a recognized summary header (substring match).

    Returns (selected_sheet_name | None, ignored_sheets_list).
    """
    visible = []
    hidden = []
    for name in wb.sheetnames:
        s = wb[name]
        # openpyxl exposes sheet_state: 'visible'|'hidden'|'veryHidden'.
        # read-only workbooks may not expose it on every sheet object;
        # log when we fall back so the audit trail records the
        # ambiguity (Codex chunk-9-step3 (h)).
        if not hasattr(s, "sheet_state"):
            log.info("xlsx_sheet_state_missing", extra={"sheet_name": name})
            visible.append(name)
            continue
        if s.sheet_state == "visible":
            visible.append(name)
        else:
            hidden.append(name)

    # Exact 'Tasks' match (case-insensitive) wins.
    for name in visible:
        if name.strip().lower() == "tasks":
            ignored = [n for n in wb.sheetnames if n != name]
            return name, ignored

    # Else first visible sheet whose header row has a recognized
    # task-summary column.
    for name in visible:
        s = wb[name]
        try:
            first_row = next(s.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            continue
        header_strs = [
            (c or "").strip().lower() if isinstance(c, str) else ""
            for c in first_row
        ]
        for hint in _SUMMARY_HINTS:
            if any(hint in h for h in header_strs):
                ignored = [n for n in wb.sheetnames if n != name]
                return name, ignored

    return None, list(wb.sheetnames)


def _cell_to_str(val: Any) -> str:
    """Render an openpyxl cell value as the canonical CSV cell string.

    Numeric-but-display-as-text dates / floats get the .isoformat() or
    str() default. None becomes empty string.
    """
    if val is None:
        return ""
    # openpyxl returns datetime / date for date-formatted cells when
    # data_only=True. ISO renders back to a parseable form.
    if hasattr(val, "isoformat"):
        try:
            return val.isoformat()
        except Exception:
            pass
    return str(val)
